"""
Payroll Export Utilities
Generate Excel payslips and payroll reports
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class PayrollExporter:
    """Export payroll data to Excel"""
    
    def __init__(self):
        # Use absolute path for reports directory
        import os
        basedir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
        self.reports_dir = os.path.join(basedir, 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_employee_payslip(self, employee_data, payroll_data):
        """
        Generate individual employee payslip
        
        Args:
            employee_data (dict): Employee information
            payroll_data (dict): Payroll computation details
            
        Returns:
            str: Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Payslip"
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Header styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        # Title
        ws['A1'] = "WORKVAULT HR & PAYROLL SYSTEM"
        ws['A1'].font = Font(bold=True, size=16, color="0066CC")
        ws.merge_cells('A1:B1')
        
        ws['A2'] = "PAYSLIP"
        ws['A2'].font = Font(bold=True, size=14)
        ws.merge_cells('A2:B2')
        
        # Employee Information
        row = 4
        ws[f'A{row}'] = "EMPLOYEE INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        info_data = [
            ("Employee Name:", employee_data.get('name', 'N/A')),
            ("Employee ID:", employee_data.get('id', 'N/A')),
            ("Department:", employee_data.get('department', 'N/A')),
            ("Position:", employee_data.get('position', 'Employee')),
            ("Pay Period:", f"{payroll_data.get('period_start', '')} to {payroll_data.get('period_end', '')}"),
            ("Payment Date:", datetime.now().strftime('%Y-%m-%d')),
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Earnings Section
        ws[f'A{row}'] = "EARNINGS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        earnings_data = [
            ("Basic Pay (Semi-Monthly)", f"₱{payroll_data.get('basic_pay', 0):,.2f}"),
            ("Overtime Pay", f"₱{payroll_data.get('overtime_pay', 0):,.2f}"),
        ]
        
        for label, value in earnings_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        # Gross Earnings
        ws[f'A{row}'] = "GROSS EARNINGS"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"₱{payroll_data.get('gross_earnings', 0):,.2f}"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 2
        
        # Deductions Section
        ws[f'A{row}'] = "DEDUCTIONS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        
        # Attendance Deductions
        ws[f'A{row}'] = "Attendance Deductions:"
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        row += 1
        
        attendance_deductions = [
            ("  Tardiness", f"₱{payroll_data.get('tardiness_deduction', 0):,.2f}"),
            ("  Undertime", f"₱{payroll_data.get('undertime_deduction', 0):,.2f}"),
            ("  Absent Days", f"₱{payroll_data.get('absent_deduction', 0):,.2f}"),
        ]
        
        for label, value in attendance_deductions:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        row += 1
        
        # Government Contributions
        ws[f'A{row}'] = "Government Contributions:"
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        row += 1
        
        gov_contributions = [
            ("  SSS Contribution", f"₱{payroll_data.get('sss_contribution', 0):,.2f}"),
            ("  PhilHealth Contribution", f"₱{payroll_data.get('philhealth_contribution', 0):,.2f}"),
            ("  Pag-IBIG Contribution", f"₱{payroll_data.get('pagibig_contribution', 0):,.2f}"),
            ("  Withholding Tax", f"₱{payroll_data.get('withholding_tax', 0):,.2f}"),
        ]
        
        for label, value in gov_contributions:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        row += 1
        
        # Other Deductions
        if payroll_data.get('other_deductions', 0) > 0:
            ws[f'A{row}'] = "Other Deductions:"
            ws[f'A{row}'].font = Font(bold=True, italic=True)
            row += 1
            
            ws[f'A{row}'] = "  Loans/Advances"
            ws[f'B{row}'] = f"₱{payroll_data.get('other_deductions', 0):,.2f}"
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
            row += 1
        
        # Total Deductions
        ws[f'A{row}'] = "TOTAL DEDUCTIONS"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"₱{payroll_data.get('total_deductions', 0):,.2f}"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 2
        
        # Net Pay
        ws[f'A{row}'] = "NET PAY"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        ws[f'B{row}'] = f"₱{payroll_data.get('net_pay', 0):,.2f}"
        ws[f'B{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'B{row}'].fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        
        row += 2
        
        # Rate Information
        ws[f'A{row}'] = "RATE INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        rate_info = [
            ("Daily Rate:", f"₱{payroll_data.get('daily_rate', 0):,.2f}"),
            ("Hourly Rate:", f"₱{payroll_data.get('hourly_rate', 0):,.2f}"),
            ("Per-Minute Rate:", f"₱{payroll_data.get('per_minute_rate', 0):,.2f}"),
        ]
        
        for label, value in rate_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(size=9)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(size=9)
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        row += 2
        
        # Footer
        ws[f'A{row}'] = "This is a computer-generated payslip. No signature required."
        ws[f'A{row}'].font = Font(italic=True, size=9, color="666666")
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Save file
        filename = f"payslip_{employee_data.get('name', 'employee').replace(' ', '_')}_{payroll_data.get('period_end', datetime.now().strftime('%Y%m%d'))}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def generate_payroll_summary(self, payroll_records):
        """
        Generate payroll summary report for all employees
        
        Args:
            payroll_records (list): List of payroll records with employee data
            
        Returns:
            str: Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Summary"
        
        # Header
        ws['A1'] = "WORKVAULT PAYROLL SUMMARY REPORT"
        ws['A1'].font = Font(bold=True, size=14, color="0066CC")
        ws.merge_cells('A1:N1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10)
        ws.merge_cells('A2:N2')
        
        # Column headers
        headers = [
            "Employee ID", "Employee Name", "Department", "Period Start", "Period End",
            "Basic Pay", "Overtime", "Gross", "Deductions", "SSS", "PhilHealth",
            "Pag-IBIG", "Tax", "Net Pay"
        ]
        
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        row = 5
        total_gross = 0
        total_deductions = 0
        total_net = 0
        
        for record in payroll_records:
            employee = record.get('users', {})
            
            ws.cell(row=row, column=1, value=record.get('user_id'))
            ws.cell(row=row, column=2, value=employee.get('name', 'N/A'))
            ws.cell(row=row, column=3, value=employee.get('department', 'N/A'))
            ws.cell(row=row, column=4, value=record.get('period_start'))
            ws.cell(row=row, column=5, value=record.get('period_end'))
            ws.cell(row=row, column=6, value=record.get('base_salary', 0))
            ws.cell(row=row, column=7, value=record.get('overtime_pay', 0))
            ws.cell(row=row, column=8, value=record.get('gross_earnings', 0))
            ws.cell(row=row, column=9, value=record.get('total_deductions', 0))
            ws.cell(row=row, column=10, value=record.get('sss_contribution', 0))
            ws.cell(row=row, column=11, value=record.get('philhealth_contribution', 0))
            ws.cell(row=row, column=12, value=record.get('pagibig_contribution', 0))
            ws.cell(row=row, column=13, value=record.get('withholding_tax', 0))
            ws.cell(row=row, column=14, value=record.get('net_pay', 0))
            
            # Format currency columns
            for col in range(6, 15):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '₱#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            
            total_gross += record.get('gross_earnings', 0)
            total_deductions += record.get('total_deductions', 0)
            total_net += record.get('net_pay', 0)
            
            row += 1
        
        # Totals row
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.merge_cells(f'A{row}:G{row}')
        
        ws.cell(row=row, column=8, value=total_gross)
        ws.cell(row=row, column=9, value=total_deductions)
        ws.cell(row=row, column=14, value=total_net)
        
        for col in [8, 9, 14]:
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.number_format = '₱#,##0.00'
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='right')
        
        # Adjust column widths
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        ws.column_dimensions['B'].width = 25  # Employee Name
        ws.column_dimensions['C'].width = 20  # Department
        
        # Save file
        filename = f"payroll_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        
        return filepath


# Convenience functions
def export_employee_payslip(employee_data, payroll_data):
    """Export single employee payslip"""
    exporter = PayrollExporter()
    return exporter.generate_employee_payslip(employee_data, payroll_data)


def export_payroll_summary(payroll_records):
    """Export payroll summary for all employees"""
    exporter = PayrollExporter()
    return exporter.generate_payroll_summary(payroll_records)
