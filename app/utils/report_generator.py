import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
import os

def generate_payroll_report(records, format_type='excel', period_start=None, period_end=None):
    """Generate payroll report in Excel or PDF format"""
    # Get the app directory (where app folder is located)
    app_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    reports_dir = os.path.join(app_dir, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if format_type == 'excel':
        filename = os.path.join(reports_dir, f'payroll_report_{timestamp}.xlsx')
        return _generate_payroll_excel(records, filename, period_start, period_end)
    else:
        filename = os.path.join(reports_dir, f'payroll_report_{timestamp}.pdf')
        return _generate_payroll_pdf(records, filename, period_start, period_end)

def _generate_payroll_excel(records, filename, period_start, period_end):
    """Generate Excel payroll report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Report"
    
    # Title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"Payroll Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Period
    if period_start and period_end:
        ws.merge_cells('A2:I2')
        period_cell = ws['A2']
        period_cell.value = f"Period: {period_start} to {period_end}"
        period_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Employee Name', 'Email', 'Department', 'Base Salary', 
               'Days Present', 'Days Absent', 'Days on Leave', 'Deductions', 'Net Salary']
    header_row = 4
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, record in enumerate(records, start=header_row + 1):
        user = record.get('users', {})
        ws.cell(row=row_idx, column=1, value=user.get('name', 'N/A'))
        ws.cell(row=row_idx, column=2, value=user.get('email', 'N/A'))
        ws.cell(row=row_idx, column=3, value=user.get('department', 'N/A'))
        ws.cell(row=row_idx, column=4, value=record.get('base_salary', 0))
        ws.cell(row=row_idx, column=5, value=record.get('days_present', 0))
        ws.cell(row=row_idx, column=6, value=record.get('days_absent', 0))
        ws.cell(row=row_idx, column=7, value=record.get('days_on_leave', 0))
        ws.cell(row=row_idx, column=8, value=record.get('deductions', 0))
        ws.cell(row=row_idx, column=9, value=record.get('net_salary', 0))
    
    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    wb.save(filename)
    return filename

def _generate_payroll_pdf(records, filename, period_start, period_end):
    """Generate PDF payroll report"""
    doc = SimpleDocTemplate(filename, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=1  # Center
    )
    
    elements.append(Paragraph("Payroll Report", title_style))
    
    if period_start and period_end:
        period_text = f"Period: {period_start} to {period_end}"
        elements.append(Paragraph(period_text, styles['Normal']))
    
    elements.append(Spacer(1, 0.3 * inch))
    
    # Table data
    data = [['Name', 'Department', 'Base Salary', 'Days Present', 
             'Days Absent', 'Deductions', 'Net Salary']]
    
    for record in records:
        user = record.get('users', {})
        data.append([
            user.get('name', 'N/A'),
            user.get('department', 'N/A'),
            f"${record.get('base_salary', 0):,.2f}",
            str(record.get('days_present', 0)),
            str(record.get('days_absent', 0)),
            f"${record.get('deductions', 0):,.2f}",
            f"${record.get('net_salary', 0):,.2f}"
        ])
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return filename


def generate_attendance_report(records, format_type='excel', start_date=None, end_date=None):
    """Generate attendance report in Excel or PDF format"""
    # Get the app directory (where app folder is located)
    app_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    reports_dir = os.path.join(app_dir, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if format_type == 'excel':
        filename = os.path.join(reports_dir, f'attendance_report_{timestamp}.xlsx')
        return _generate_attendance_excel(records, filename, start_date, end_date)
    else:
        filename = os.path.join(reports_dir, f'attendance_report_{timestamp}.pdf')
        return _generate_attendance_pdf(records, filename, start_date, end_date)

def _generate_attendance_excel(records, filename, start_date, end_date):
    """Generate Excel attendance report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "Attendance Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Period
    if start_date and end_date:
        ws.merge_cells('A2:G2')
        period_cell = ws['A2']
        period_cell.value = f"Period: {start_date} to {end_date}"
        period_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Employee Name', 'Email', 'Department', 'Date', 'Time In', 'Time Out', 'Status']
    header_row = 4
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, record in enumerate(records, start=header_row + 1):
        user = record.get('users', {})
        ws.cell(row=row_idx, column=1, value=user.get('name', 'N/A'))
        ws.cell(row=row_idx, column=2, value=user.get('email', 'N/A'))
        ws.cell(row=row_idx, column=3, value=user.get('department', 'N/A'))
        ws.cell(row=row_idx, column=4, value=record.get('date', 'N/A'))
        ws.cell(row=row_idx, column=5, value=record.get('time_in', 'N/A'))
        ws.cell(row=row_idx, column=6, value=record.get('time_out', 'N/A'))
        ws.cell(row=row_idx, column=7, value=record.get('status', 'N/A'))
    
    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    wb.save(filename)
    return filename

def _generate_attendance_pdf(records, filename, start_date, end_date):
    """Generate PDF attendance report"""
    doc = SimpleDocTemplate(filename, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=1
    )
    
    elements.append(Paragraph("Attendance Report", title_style))
    
    if start_date and end_date:
        period_text = f"Period: {start_date} to {end_date}"
        elements.append(Paragraph(period_text, styles['Normal']))
    
    elements.append(Spacer(1, 0.3 * inch))
    
    # Table data
    data = [['Name', 'Department', 'Date', 'Time In', 'Time Out', 'Status']]
    
    for record in records:
        user = record.get('users', {})
        data.append([
            user.get('name', 'N/A'),
            user.get('department', 'N/A'),
            record.get('date', 'N/A'),
            record.get('time_in', 'N/A'),
            record.get('time_out', 'N/A') or '-',
            record.get('status', 'N/A')
        ])
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return filename
